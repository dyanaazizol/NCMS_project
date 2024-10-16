from django.shortcuts import render, redirect, get_object_or_404, HttpResponseRedirect
from django.contrib import messages
from django.contrib.messages import get_messages
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from .models import  Division, Doer, DCC, Scenario, BGCM, HCBD, Level, Action, Rate, NCReport, Admin
from django.db.models import Q
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def login(request):
    if request.method == 'POST':
        # Doer Login
        if 'd_doerID' in request.POST:
            doerID = request.POST['d_doerID']
            doerPass = request.POST['d_doerPass']

            if doerID == "doer" and doerPass == "doer":
                return redirect('doerHomepage')
            else:
                try:
                    doer = Doer.objects.get(doer_id=doerID)
                    if doer.password == doerPass:  # Kalau data betul n sama
                        request.session['doerID'] = doerID
                        request.session['doerName'] = doer.name
                        return redirect('doerHomepage')
                    else:
                        return render(request, 'login.html', {'error_message': 'Incorrect ID or Password'})
                except Doer.DoesNotExist:
                    return render(request, 'login.html', {'error_message': 'No result found'})

        # DCC Login
        elif 'd_dccID' in request.POST:
            dccID = request.POST['d_dccID']
            dccPass = request.POST['d_dccPass']

            if dccID == "dcc" and dccPass == "dcc":
                return redirect('dccHomepage')
            else:
                try:
                    dcc = DCC.objects.get(dcc_id=dccID)
                    if dcc.password == dccPass:  # Kalau data betul n sama
                        request.session['dccID'] = dccID
                        request.session['dccName'] = dcc.name
                        request.session['dccDivision'] = dcc.division.name
                        return redirect('dccHomepage')
                    else:
                        return render(request, 'login.html', {'error_message': 'Incorrect ID or Password'})
                except DCC.DoesNotExist:
                    return render(request, 'login.html', {'error_message': 'No result found'})

        # BGCM Login
        elif 'b_bgcmID' in request.POST:
            bgcmID = request.POST['b_bgcmID']
            bgcmPass = request.POST['b_bgcmPass']

            try:
                bgcm = BGCM.objects.get(bgcm_id=bgcmID)
                if bgcm.password == bgcmPass: # Kalau data betul n sama
                    request.session['bgcmID'] = bgcmID
                    request.session['bgcmName'] = bgcm.name
                    return redirect('BGCMHomepage')
                else:
                    return render(request, 'login.html', {'error_message': 'Incorrect ID or Password'})
            except BGCM.DoesNotExist:
                return render(request, 'login.html', {'error_message': 'No result found'})

        # HCBD Login
        elif 'h_hcbdID' in request.POST:
            hcbdID = request.POST['h_hcbdID']
            hcbdPass = request.POST['h_hcbdPass']

            try:
                hcbd = HCBD.objects.get(hcbd_id=hcbdID)
                if hcbd.password == hcbdPass: # Kalau data betul n sama
                    request.session['hcbdID'] = hcbdID
                    request.session['hcbdName'] = hcbd.name
                    return redirect('HCBDHomepage')
                else:
                    return render(request, 'login.html', {'error_message': 'Incorrect ID or Password'})
            except HCBD.DoesNotExist:
                return render(request, 'login.html', {'error_message': 'No result found'})
            
        # ADMIN Login
        elif 'a_adminID' in request.POST:
            adminID = request.POST['a_adminID']
            adminPass = request.POST['a_adminPass']

            try:
                admin = Admin.objects.get(admin_id=adminID)
                if admin.password == adminPass: # Kalau data betul n sama
                    request.session['adminID'] = adminID
                    request.session['adminName'] = admin.admin_name
                    return redirect('AdminHomepage')
                else:
                    return render(request, 'login.html', {'error_message': 'Incorrect ID or Password'})
            except Admin.DoesNotExist:
                return render(request, 'login.html', {'error_message': 'No result found'})

    return render(request, 'login.html')

def doerHomepage(request):
    return render(request, 'doerHomepage.html')

def dccHomepage(request):
    return render(request, 'dccHomepage.html')

def BGCMHomepage(request):
    return render(request, 'BGCMHomepage.html')

def HCBDHomepage(request):
    return render(request, 'HCBDHomepage.html')

################################### DCC #################################################################

def dccAdddoer(request):
    d_divisionID = Division.objects.all()

    if request.method == "POST":
        doID = request.POST['d_doerID']
        doName = request.POST['d_doerName']
        doPass = request.POST['d_doerPass']
        doEmail = request.POST['d_doerEmail']
        doDiv = request.POST['d_doerDivision']
        doPosition = request.POST['d_doerPosition']
        doDirectsupervisor = request.POST['d_directSupervisor']
        doGMHOD = request.POST['d_doerGmHOD']
        doUnit = request.POST['d_doerUnit']
        doState = request.POST['d_doerState']
        doSubsidiary = request.POST['d_doerSubsidiary']
        doLocation = request.POST['d_doerlocation']
        doVerticalsegment = request.POST['d_doerVerticalSeg']

        # Check if Doer with same ID already exists
        if Doer.objects.filter(doer_id=doID).exists():
            dict = {
                'd_divisionID': d_divisionID,
                'message': "Doer with this ID already exists"  # Error message
            }
            return render(request, 'dccAdddoer.html', dict)

        # Get the foreign key reference for division
        doer_division = Division.objects.get(division_id=doDiv)

        # Create a new Doer entry
        data = Doer(
            doer_id=doID, name=doName, password=doPass, doer_email=doEmail, division=doer_division, 
            position=doPosition, direct_supervisor=doDirectsupervisor, gm_hod=doGMHOD, 
            unit=doUnit, state=doState, subsidiary=doSubsidiary, 
            location=doLocation, vertical=doVerticalsegment
        )

        # Save the new Doer
        data.save()

        dict = {
            'd_divisionID': d_divisionID,
            'message': "Doer successfully added"  # Success message
        }
        return render(request, 'dccAdddoer.html', dict)

    else:
        dict = {
            'd_divisionID': d_divisionID,
        }
    return render(request, 'dccAdddoer.html', dict)

def dccAction(request):
    return render(request, 'dccAction.html')

def dccChooseNC(request):
    return render(request, 'dccChooseNC.html')

def dccCreateNC(request):
    if 'dccID' in request.session:
        dc_ID = request.session.get('dccID')  

        try:
            # based on dcc login
            dc = DCC.objects.get(dcc_id=dc_ID)
        except DCC.DoesNotExist:
            return redirect('login')  

        d_doID = Doer.objects.all()
        s_scenID = Scenario.objects.all()
        n_ncRate = Rate.objects.all()

        context = {
            'dcc_division': dc.division,  
            'd_doID': d_doID,
            's_scenID': s_scenID,
            'n_ncRate':n_ncRate
        }

        if request.method == "POST":
            doID = request.POST.get('d_doerID')
            scenID = request.POST.get('s_scenarioID')
            policyName = request.POST.get('p_policyName')
            processOwner = request.POST.get('p_po')
            dateIncident = request.POST.get('d_dateIncident')
            typeRef = request.POST.get('t_typeRef')
            refNo = request.POST.get('r_referenceNo')
            conProjectName = request.POST.get('p_projectName')
            accName = request.POST.get('a_accountName')
            poDivision = request.POST.get('p_poDivision')
            potentialFi = request.POST.get('p_potFi')
            nonfinancialImpact = request.POST.get('n_nonFi')
            frequency = request.POST.get('f_freq')
            ncRating = request.POST.get('n_ncRating')
            levelImpact = request.POST.get('l_levImpact')
            act = request.POST.get('a_action')
            scenarioDetails = request.POST.get('s_scenarioDetails')

            doer = get_object_or_404(Doer, doer_id=doID)
            scenario = get_object_or_404(Scenario, scenario_id=scenID)
            ratee = get_object_or_404(Rate, rate_id=ncRating)

            # Create NC report
            nc_report = NCReport(
                dccID=dc,  
                doerID=doer,
                scenarioID=scenario,
                policy_name=policyName,
                process_owner=processOwner,
                dateIncident=dateIncident,
                typeRef=typeRef,
                refNo=refNo,
                conProjectName=conProjectName,
                accName=accName,
                poDivision=poDivision,
                poFI=potentialFi,
                nonFI=nonfinancialImpact,
                frequency=frequency,
                rateID=ratee,
                level=levelImpact,
                action=act,
                scenarioDetails=scenarioDetails,
            )
            nc_report.save()

            context['message'] = ""
            context['nc_report'] = nc_report  

            return render(request, 'dccCreateNC.html', context)

        return render(request, 'dccCreateNC.html', context)
    else:
        return redirect('login')


    
def dccSearchNC(request):
    status_filter = request.GET.get('status')
    
    # dcc based on login
    dcc_id = request.session.get('dccID')
    
    # display dcc based on login
    ncreports = NCReport.objects.filter(dccID__dcc_id=dcc_id)

    # Update NCReport status
    for nc_report in ncreports:
        if nc_report.doerJustification == 'pending..':
            nc_report.status = 'pending doer justification..'
        elif nc_report.remarksBGCM == 'pending..':
            nc_report.status = 'pending remarks BGCM..'
        elif nc_report.remarksHCBD == 'pending..':
            nc_report.status = 'pending remarks HCBD..'
        elif nc_report.ncDecision == 'pending..':
            nc_report.status = 'pending decision..'
        elif nc_report.acknowledgment == 'pending..':
            nc_report.status = 'pending doer acknowledgment..'
        else:
            nc_report.status = 'completed'

        # Save updated status
        nc_report.save()

    # Filter NC reports based on the selected status
    if status_filter == 'pending_justification':
        ncreports = ncreports.filter(doerJustification='pending..')
    elif status_filter == 'pending_bgcm':
        ncreports = ncreports.filter(remarksBGCM='pending..')
    elif status_filter == 'pending_hcbd':
        ncreports = ncreports.filter(remarksHCBD='pending..')
    elif status_filter == 'pending_decision':
        ncreports = ncreports.filter(ncDecision='pending..')
    elif status_filter == 'pending_acknowledgment':
        ncreports = ncreports.filter(acknowledgment='pending..')
    elif status_filter == 'completed':
        # Filter completed reports (exclude reports with any field still pending)
        ncreports = ncreports.exclude(
            Q(doerJustification='pending..') |
            Q(remarksBGCM='pending..') |
            Q(remarksHCBD='pending..') |
            Q(ncDecision='pending..') |
            Q(acknowledgment='pending..')
        )

    return render(request, 'dccSearchNC.html', {
        'list_nc': ncreports,
    })



def delete_Report(request, report_id):
        report = get_object_or_404(NCReport, pk=report_id)
        report.delete()
        return JsonResponse({'success': True})

def get_scenario_details(request):
    scenario_id = request.GET.get('scenario_id')
    if scenario_id:
        try:
            scenario = Scenario.objects.get(scenario_id=scenario_id)
            response_data = {
                'policy_name': scenario.policy.name,
                'process_owner': scenario.policy.process_owner.name,
            }
            return JsonResponse(response_data)
        except Scenario.DoesNotExist:
            return JsonResponse({'error': 'Scenario not found'}, status=404)
    return JsonResponse({'error': 'Invalid request'}, status=400)

def get_ncRating_details(request):
    rate_id = request.GET.get('rate_id')
    if rate_id:
        try:
            rate = Rate.objects.get(rate_id=rate_id)
            response_data = {
                'level': rate.actionID.levelID.level_id, 
                'action': rate.actionID.action_details 
            }
            return JsonResponse(response_data)
        except Rate.DoesNotExist:
            return JsonResponse({'error': 'Rate not found'}, status=404)
    return JsonResponse({'error': 'Invalid request'}, status=400)

def dccViewNC(request, report_id):
    report = get_object_or_404(NCReport, id=report_id)

    # Clear previous messages
    storage = get_messages(request)
    for _ in storage:
        pass
    
    if request.method == 'POST':
        new_clarificationDate = request.POST.get('c_clarificationDate')
        new_ncDecision = request.POST.get('n_ncDecision')
        new_remarksPO = request.POST.get('r_remarksPo')
        report.clarificationDate = new_clarificationDate
        report.ncDecision = new_ncDecision
        report.remarksPO = new_remarksPO
        report.save()

        # Add a success message
        messages.success(request, 'updated successfully.')
        return redirect('dccViewNC', report_id=report_id)

    context = {
        'report_id': report_id,
        'status': report.status,
        'doer_id': report.doerID.doer_id,
        'name': report.doerID.name,
        'doer_email': report.doerID.doer_email,
        'division': report.doerID.division.name,  
        'position': report.doerID.position,
        'direct_supervisor': report.doerID.direct_supervisor,
        'gm_hod': report.doerID.gm_hod,
        'unit': report.doerID.unit,
        'state': report.doerID.state,
        'subsidiary': report.doerID.subsidiary,
        'location': report.doerID.location,
        'vertical': report.doerID.vertical,
        'scenarioID': report.scenarioID.scenario_id,
        'policy_name': report.policy_name,
        'process_owner' : report.process_owner,
        'dateIncident': report.dateIncident,
        'typeRef': report.typeRef,
        'refNo': report.refNo,
        'conProjectName': report.conProjectName,
        'accName': report.accName,
        'poDivision': report.poDivision,
        'scenarioDetails': report.scenarioDetails,
        'poFI': report.poFI,
        'nonFI': report.nonFI,
        'frequency': report.frequency,
        'ncRating': report.rateID.rate_id,
        'levelImpact': report.level,
        'doerJustification': report.doerJustification,
        'remarksBGCM': report.remarksBGCM,
        'remarksHCBD': report.remarksHCBD,
        'clarificationDate': report.clarificationDate,
        'ncDecision': report.ncDecision,
        'remarksPO': report.remarksPO,
        'action': report.action,
        'acknowledgment': report.acknowledgment,
    }
    return render(request, 'dccViewNC.html', context)

def exportExcel(request):
    # nak fetch daripada models
    qs = NCReport.objects.all()

    # Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="NC_Report.xlsx"'

    # Create workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'NC Report'

    # column headers
    columns = [
        'NC ID', 'Doer ID', 'Doer Name', 'Division', 'Doer Direct Supervisor', 'Position', 'Doer GM/Hod', 'Unit', 
        'State', 'Subsidiary', 'Location', 'Vertical', 'Scenario', 'Policies & Procedures', 'Process Owner', 
        'Date of Incident', 'Type of Reference No', 'Reference No', 'Contract Name / Project Name', 'Account Name', 
        'Process Owner (Div)', 'NC Scenario Details', 'Potential Financial Impact Value', 'Non-Financial Impact', 
        'Frequency', 'NC Rating', 'Level of Impact', 'Consequence Action to Doer', 'Justification', 
        'Remarks BGCM', 'Remarks HCBD', 'Clarification Date', 'Clarification Decision', 'Remarks Process Owner', 
        'Acknowledgement', 'Status'
    ]
    
    # Header row
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="50C878")
        cell.font = Font(bold=True, color="F7F6FA")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # data based on header
    for report in qs:
        row_num += 1
        row = [
            f"{report.dccID.division.name}/{report.id:04}/{report.dateIncident.year}",  
            report.doerID_id,
            report.doerID.name,
            report.doerID.division.name,
            report.doerID.position,
            report.doerID.direct_supervisor,
            report.doerID.gm_hod,
            report.doerID.unit,
            report.doerID.state,
            report.doerID.subsidiary,
            report.doerID.location,
            report.doerID.vertical,
            report.scenarioID.name,
            report.policy_name,
            report.process_owner,
            report.dateIncident,
            report.typeRef,
            report.refNo,
            report.conProjectName,
            report.accName,
            report.poDivision,
            report.scenarioDetails,
            report.poFI,
            report.nonFI,
            report.frequency,
            report.rateID.rate_id,
            report.level,
            report.action,
            report.doerJustification,
            report.remarksBGCM,
            report.remarksHCBD,
            report.clarificationDate,
            report.ncDecision,
            report.remarksPO,
            report.acknowledgment,
            report.status
        ]
        
        # setiap cell in a row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # delete default sheet
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Save workbook
    workbook.save(response)

    return response

def dccProfile(request):
    if 'dccID' in request.session:
        dccid = request.session['dccID']
        try:
            dcc = DCC.objects.get(dcc_id = dccid)
        except DCC.DoesNotExist:
            return redirect('login')
        
        context = {
            'dcc' : dcc
        }
        return render(request, 'dccProfile.html', context)
    else:
        return redirect('login')
    
################################### doer #################################################################

def doerProfile(request):
    if 'doerID' in request.session:
        doerid = request.session['doerID']
        try:
            doer = Doer.objects.get(doer_id = doerid)
        except Doer.DoesNotExist:
            return redirect('login')
        
        context = {
            'doer' : doer
        }
        return render(request, 'doerProfile.html', context)
    else:
        return redirect('login')
    
def doerSearchNC(request):
    doer_id = request.session.get('doerID')  # based on login
    ncreports = None 
    dcc_division = 'Unknown'  

    if doer_id:
        all_reports = NCReport.objects.filter(doerID__doer_id=doer_id)
        ncreports = all_reports

        for nc_report in ncreports:
            if nc_report.doerJustification == 'pending..':
                nc_report.status = 'pending doer justification..'
            elif nc_report.remarksBGCM == 'pending..':
                nc_report.status = 'pending remarks BGCM..'
            elif nc_report.remarksHCBD == 'pending..':
                nc_report.status = 'pending remarks HCBD..'
            elif nc_report.ncDecision == 'pending..':
                nc_report.status = 'pending Decision'
            elif nc_report.acknowledgment == 'pending..':
                nc_report.status = 'pending doer acknowledgement..'
            else:
                nc_report.status = 'completed'

            # Save
            nc_report.save()

        if ncreports.exists():
            first_report = ncreports.first()
            dcc_id = first_report.dccID.dcc_id

            dcc_division_obj = DCC.objects.filter(dcc_id=dcc_id).first()
            if dcc_division_obj:
                dcc_division = dcc_division_obj.division.name

    return render(request, 'doerSearchNC.html', {
        'list_nc': ncreports,
        'report': ncreports,
        'dcc_division': dcc_division
    })


def approve_acknowledgement(request, report_id):
    report = get_object_or_404(NCReport, id=report_id)
    
    if request.method == 'POST':
        acknowledge_status = request.POST.get('acknowledge_status')
        
        if acknowledge_status and acknowledge_status == 'yes':
            report.acknowledgment = acknowledge_status
            report.save()
            return JsonResponse({'success': True})
        
        return JsonResponse({'success': False, 'message': 'Invalid acknowledgment status'}, status=400)
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

def doerViewNC(request, report_id):
    report = get_object_or_404(NCReport, id=report_id)
    
    # Clear previous messages
    storage = get_messages(request)
    for _ in storage:
        pass

    if request.method == 'POST':
        new_justification = request.POST.get('j_just')
        report.doerJustification = new_justification
        report.save()
        
        # Add a new success message
        messages.success(request, 'updated successfully.')
        return redirect('doerViewNC', report_id=report_id)

    context = {
        'report_id': report_id,
        'doer_id': report.doerID.doer_id,
        'name': report.doerID.name,
        'doer_email': report.doerID.doer_email,
        'division': report.doerID.division.name,
        'position': report.doerID.position,
        'direct_supervisor': report.doerID.direct_supervisor,
        'gm_hod': report.doerID.gm_hod,
        'unit': report.doerID.unit,
        'state': report.doerID.state,
        'subsidiary': report.doerID.subsidiary,
        'location': report.doerID.location,
        'vertical': report.doerID.vertical,
        'scenarioID': report.scenarioID.scenario_id,
        'policy_name': report.policy_name,
        'process_owner': report.process_owner,
        'dateIncident': report.dateIncident,
        'typeRef': report.typeRef,
        'refNo': report.refNo,
        'conProjectName': report.conProjectName,
        'accName': report.accName,
        'poDivision': report.poDivision,
        'scenarioDetails': report.scenarioDetails,
        'poFI': report.poFI,
        'nonFI': report.nonFI,
        'frequency': report.frequency,
        'ncRating': report.rateID.rate_id,
        'levelImpact': report.level,
        'doerJustification': report.doerJustification,
        'remarksBGCM': report.remarksBGCM,
        'remarksHCBD': report.remarksHCBD,
        'clarificationDate': report.clarificationDate,
        'ncDecision': report.ncDecision,
        'remarksPO': report.remarksPO,
        'action': report.action,
        'acknowledgment': report.acknowledgment,
        'status': report.status,
    }
    return render(request, 'doerViewNC.html', context)

################################### BGCM #################################################################

def BGCMSearchNC(request):
    status_filter = request.GET.get('status')
    ncreports = NCReport.objects.all()

    for nc_report in ncreports:
        if nc_report.doerJustification == 'pending..':
            nc_report.status = 'pending doer justification..'
        elif nc_report.remarksBGCM == 'pending..':
            nc_report.status = 'pending remarks BGCM..'
        elif nc_report.remarksHCBD == 'pending..':
            nc_report.status = 'pending remarks HCBD..'
        elif nc_report.ncDecision == 'pending..':
            nc_report.status = 'pending decision..'
        elif nc_report.acknowledgment == 'pending..':
            nc_report.status = 'pending doer acknowledgment..'
        else:
            nc_report.status = 'completed'

        # Save
        nc_report.save()

    # Apply filters based on the selected status
    if status_filter == 'pending_justification':
        ncreports = ncreports.filter(doerJustification='pending..')
    elif status_filter == 'pending_bgcm':
        ncreports = ncreports.filter(remarksBGCM='pending..')
    elif status_filter == 'pending_hcbd':
        ncreports = ncreports.filter(remarksHCBD='pending..')
    elif status_filter == 'pending_decision':
        ncreports = ncreports.filter(ncDecision='pending..')
    elif status_filter == 'pending_acknowledgment':
        ncreports = ncreports.filter(acknowledgment='pending..')
    elif status_filter == 'completed':
        # Filter completed reports (exclude reports with any field still pending)
        ncreports = ncreports.exclude(
            Q(doerJustification='pending..') |
            Q(remarksBGCM='pending..') |
            Q(remarksHCBD='pending..') |
            Q(ncDecision='pending..') |
            Q(acknowledgment='pending..')
        ) 

    return render(request, 'BGCMSearchNC.html', {
        'list_nc': ncreports,
    })

def BGCMViewNC(request, report_id):
    report = get_object_or_404(NCReport, id=report_id)

    # Clear previous messages
    storage = get_messages(request)
    for _ in storage:
        pass
    
    if request.method == 'POST':
        new_remarksbgcm = request.POST.get('r_remarksBgcm')
        report.remarksBGCM = new_remarksbgcm
        report.save()

        # Add a success message
        messages.success(request, 'updated successfully.')
    
        # Redirect to the same page to prevent form resubmission on refresh
        return redirect('BGCMViewNC', report_id=report_id)
    
    context = {
        'report_id': report_id,
        'status': report.status, 
        'doer_id': report.doerID.doer_id,
        'name': report.doerID.name,
        'doer_email': report.doerID.doer_email,
        'division': report.doerID.division.name,  
        'position': report.doerID.position,
        'direct_supervisor': report.doerID.direct_supervisor,
        'gm_hod': report.doerID.gm_hod,
        'unit': report.doerID.unit,
        'state': report.doerID.state,
        'subsidiary': report.doerID.subsidiary,
        'location': report.doerID.location,
        'vertical': report.doerID.vertical,
        'scenarioID': report.scenarioID.scenario_id,
        'policy_name': report.policy_name,
        'process_owner' : report.process_owner,
        'dateIncident': report.dateIncident,
        'typeRef': report.typeRef,
        'refNo': report.refNo,
        'conProjectName': report.conProjectName,
        'accName': report.accName,
        'poDivision': report.poDivision,
        'scenarioDetails': report.scenarioDetails,
        'poFI': report.poFI,
        'nonFI': report.nonFI,
        'frequency': report.frequency,
        'ncRating': report.rateID.rate_id,
        'levelImpact': report.level,
        'doerJustification': report.doerJustification,
        'remarksBGCM': report.remarksBGCM,
        'remarksHCBD': report.remarksHCBD,
        'clarificationDate': report.clarificationDate,
        'ncDecision': report.ncDecision,
        'remarksPO': report.remarksPO,
        'action': report.action,
        'acknowledgment': report.acknowledgment,
    }
    return render(request, 'BGCMViewNC.html', context)

def BGCMProfile(request):
    if 'bgcmID' in request.session:
        bgcmid = request.session['bgcmID']
        try:
            bgcm = BGCM.objects.get(bgcm_id = bgcmid)
        except BGCM.DoesNotExist:
            return redirect('login')
        
        context = {
            'bgcm' : bgcm
        }
        return render(request, 'BGCMProfile.html', context)
    else:
        return redirect('login')
    
################################### HCBD #################################################################

def HCBDSearchNC(request):
    status_filter = request.GET.get('status')
    
    # Fetch all NCReport records
    ncreports = NCReport.objects.all()

    # Iterate through each NCReport to update the status
    for nc_report in ncreports:
        if nc_report.doerJustification == 'pending..':
            nc_report.status = 'pending doer justification..'
        elif nc_report.remarksBGCM == 'pending..':
            nc_report.status = 'pending remarks BGCM..'
        elif nc_report.remarksHCBD == 'pending..':
            nc_report.status = 'pending remarks HCBD..'
        elif nc_report.ncDecision == 'pending..':
            nc_report.status = 'pending decision..'
        elif nc_report.acknowledgment == 'pending..':
            nc_report.status = 'pending doer acknowledgment..'
        else:
            nc_report.status = 'completed'

        # Save the updated status to the database
        nc_report.save()

    # Apply filters based on the selected status
    if status_filter == 'pending_justification':
        ncreports = ncreports.filter(doerJustification='pending..')
    elif status_filter == 'pending_bgcm':
        ncreports = ncreports.filter(remarksBGCM='pending..')
    elif status_filter == 'pending_hcbd':
        ncreports = ncreports.filter(remarksHCBD='pending..')
    elif status_filter == 'pending_decision':
        ncreports = ncreports.filter(ncDecision='pending..')
    elif status_filter == 'pending_acknowledgment':
        ncreports = ncreports.filter(acknowledgment='pending..')
    elif status_filter == 'completed':
        # Filter completed reports (exclude reports with any field still pending)
        ncreports = ncreports.exclude(
            Q(doerJustification='pending..') |
            Q(remarksBGCM='pending..') |
            Q(remarksHCBD='pending..') |
            Q(ncDecision='pending..') |
            Q(acknowledgment='pending..')
        )

    return render(request, 'HCBDSearchNC.html', {
        'list_nc': ncreports,
    })

def HCBDViewNC(request, report_id):
    report = get_object_or_404(NCReport, id=report_id)
    
    if request.method == 'POST':
        new_remarkshcbd = request.POST.get('r_remarksHcbd')
        report.remarksHCBD = new_remarkshcbd
        report.save()

        # Add a success message
        messages.success(request, 'updated successfully.')

        # Redirect to the same page to prevent form resubmission on refresh
        return redirect('HCBDViewNC', report_id=report_id)

    context = {
        'report_id': report_id,
        'status': report.status,
        'doer_id': report.doerID.doer_id,
        'name': report.doerID.name,
        'doer_email': report.doerID.doer_email,
        'division': report.doerID.division.name,  
        'position': report.doerID.position,
        'direct_supervisor': report.doerID.direct_supervisor,
        'gm_hod': report.doerID.gm_hod,
        'unit': report.doerID.unit,
        'state': report.doerID.state,
        'subsidiary': report.doerID.subsidiary,
        'location': report.doerID.location,
        'vertical': report.doerID.vertical,
        'scenarioID': report.scenarioID.scenario_id,
        'policy_name': report.policy_name,
        'process_owner' : report.process_owner,
        'dateIncident': report.dateIncident,
        'typeRef': report.typeRef,
        'refNo': report.refNo,
        'conProjectName': report.conProjectName,
        'accName': report.accName,
        'poDivision': report.poDivision,
        'scenarioDetails': report.scenarioDetails,
        'poFI': report.poFI,
        'nonFI': report.nonFI,
        'frequency': report.frequency,
        'ncRating': report.rateID.rate_id,
        'levelImpact': report.level,
        'doerJustification': report.doerJustification,
        'remarksBGCM': report.remarksBGCM,
        'remarksHCBD': report.remarksHCBD,
        'clarificationDate': report.clarificationDate,
        'ncDecision': report.ncDecision,
        'remarksPO': report.remarksPO,
        'action': report.action,
        'acknowledgment': report.acknowledgment,
    }
    return render(request, 'HCBDViewNC.html', context)

def HCBDProfile(request):
    if 'hcbdID' in request.session:
        hcbdid = request.session['hcbdID']
        try:
            hcbd = HCBD.objects.get(hcbd_id = hcbdid)
        except HCBD.DoesNotExist:
            return redirect('login')
        
        context = {
            'hcbd' : hcbd
        }
        return render(request, 'HCBDProfile.html', context)
    else:
        return redirect('login')

################################### Admin #################################################################

def AdminHomepage(request):
    return render(request, 'AdminHomepage.html')

def AdminManage(request):
    return render(request, 'AdminManage.html')

def AdminManageDCC(request):
    return render(request, 'AdminManageDCC.html')

def AdminaddDCC(request):
    d_divisionID = Division.objects.all()

    if request.method == "POST":
        dccID = request.POST['d_dccID']
        dccName = request.POST['d_dccName']
        dccPass = request.POST['d_dccPass']
        dccEmail = request.POST['d_dccEmail']
        dccDiv = request.POST['d_dccDivision']

        # Check if DCC with same ID already exists
        if DCC.objects.filter(dcc_id=dccID).exists():
            dict = {
                'd_divisionID': d_divisionID,
                'message': "DCC with this ID already exists"  # Error message
            }
            return render(request, 'AdminaddDCC.html', dict)

        # Get the foreign key reference for division
        dcc_division = Division.objects.get(division_id=dccDiv)

        # Create a new Doer entry
        data = DCC(
            dcc_id=dccID, name=dccName, password=dccPass, dcc_email=dccEmail, division=dcc_division)

        # Save the new Doer
        data.save()

        dict = {
            'd_divisionID': d_divisionID,
            'message': "DCC successfully added"  # Success message
        }
        return render(request, 'AdminaddDCC.html', dict)

    else:
        dict = {
            'd_divisionID': d_divisionID,
        }
    return render(request, 'AdminaddDCC.html', dict)

def AdminSearchDCC(request):
    dcc_id = request.GET.get('dcc_id')
    
    # Fetch all DCC records for the dropdown
    dcc_list = DCC.objects.all()

    # Initialize the query to fetch all DCCs or filter based on the selected DCC ID
    if dcc_id:
        dccs = DCC.objects.filter(dcc_id=dcc_id)
    else:
        dccs = DCC.objects.all()  # Show all DCCs if no filter is applied

    return render(request, 'AdminSearchDCC.html', {
        'dcc_list': dcc_list,  # List for dropdown
        'list_dcc': dccs,      # Filtered or all DCC records to display
    })

def AdminDeleteDCC(request, dcc_id):
    dcc = get_object_or_404(DCC, dcc_id=dcc_id)  
    dcc.delete()  
    return redirect('AdminSearchDCC')  

def AdminUpdateDCC(request, dcc_id):
    dcc = get_object_or_404(DCC, dcc_id=dcc_id)
    divisions = Division.objects.all()  # Fetch all divisions from the database

    if request.method == 'POST':
        new_division_id = request.POST.get('dcc_division')
        new_division = Division.objects.get(division_id=new_division_id)  # Fetch the selected division
        dcc.division = new_division
        dcc.save()

        # Add a success message
        messages.success(request, 'Updated successfully.')

        # Redirect to the same page to prevent form resubmission on refresh
        return redirect('AdminUpdateDCC', dcc_id=dcc_id)

    context = {
        'dcc_id': dcc.dcc_id,
        'name': dcc.name,
        'password': dcc.password,
        'dcc_email': dcc.dcc_email,
        'division': dcc.division,
        'divisions': divisions,  # Pass the divisions to the template
    }
    return render(request, 'AdminUpdateDCC.html', context)

def AdminManageBGCM(request):
    return render(request, 'AdminManageBGCM.html')

def AdminaddBGCM(request):
    d_divisionID = Division.objects.all()

    if request.method == "POST":
        bgcmID = request.POST['b_bgcmID']
        bgcmName = request.POST['b_bgcmName']
        bgcmPass = request.POST['b_bgcmPass']
        bgcmEmail = request.POST['b_bgcmEmail']
        bgcmDiv = request.POST['b_bgcmDivision']

        # Check if DCC with same ID already exists
        if BGCM.objects.filter(bgcm_id=bgcmID).exists():
            dict = {
                'd_divisionID': d_divisionID,
                'message': "BGCM with this ID already exists"  # Error message
            }
            return render(request, 'AdminaddDCC.html', dict)

        # Get the foreign key reference for division
        bg_division = Division.objects.get(division_id=bgcmDiv)

        # Create a new Doer entry
        data = BGCM(
            bgcm_id=bgcmID, name=bgcmName, password=bgcmPass, bgcm_email=bgcmEmail, bgcm_division= bg_division)

        # Save the new Doer
        data.save()

        dict = {
            'd_divisionID': d_divisionID,
            'message': "BGCM successfully added"  # Success message
        }
        return render(request, 'AdminaddBGCM.html', dict)

    else:
        dict = {
            'd_divisionID': d_divisionID,
        }
    return render(request, 'AdminaddBGCM.html', dict)

def AdminSearchBGCM(request):
    bgcm_id = request.GET.get('bgcm_id')
    
    # Fetch all bgcm records for the dropdown
    bgcm_list = BGCM.objects.all()

    # Initialize the query to fetch all BGCMs or filter based on the selected BGCM ID
    if bgcm_id:
        bgcms = BGCM.objects.filter(bgcm_id=bgcm_id)
    else:
        bgcms = BGCM.objects.all()  # Show all BGCMs if no filter is applied

    return render(request, 'AdminSearchBGCM.html', {
        'bgcm_list': bgcm_list,  # List for dropdown
        'list_bgcm': bgcms,      # Filtered or all BGCM records to display
    })

def AdminDeleteBGCM(request, bgcm_id):
    bgcm = get_object_or_404(BGCM, bgcm_id=bgcm_id)  
    bgcm.delete()  
    return redirect('AdminSearchBGCM')

def AdminUpdateBGCM(request, bgcm_id):
    bgcm = get_object_or_404(BGCM, bgcm_id=bgcm_id)
    divisions = Division.objects.all()  # Fetch all divisions from the database

    if request.method == 'POST':
        new_division_id = request.POST.get('bgcm_division')
        new_division = Division.objects.get(division_id=new_division_id)  # Fetch the selected division
        bgcm.bgcm_division = new_division
        bgcm.save()

        # Add a success message
        messages.success(request, 'Updated successfully.')

        # Redirect to the same page to prevent form resubmission on refresh
        return redirect('AdminUpdateBGCM', bgcm_id=bgcm_id)

    context = {
        'bgcm_id': bgcm.bgcm_id,
        'name': bgcm.name,
        'password': bgcm.password,
        'bgcm_email': bgcm.bgcm_email,
        'division': bgcm.bgcm_division,
        'divisions': divisions,  # Pass the divisions to the template
    }
    return render(request, 'AdminUpdateBGCM.html', context)

def AdminManageHCBD(request):
    return render(request, 'AdminManageHCBD.html')

def AdminaddHCBD(request):
    d_divisionID = Division.objects.all()

    if request.method == "POST":
        hcbdID = request.POST['h_hcbdID']
        hcbdName = request.POST['h_hcbdName']
        hcbdPass = request.POST['h_hcbdPass']
        hcbdEmail = request.POST['h_hcbdEmail']
        hcbdDiv = request.POST['h_hcbdDivision']

        # Check if DCC with same ID already exists
        if HCBD.objects.filter(hcbd_id=hcbdID).exists():
            dict = {
                'd_divisionID': d_divisionID,
                'message': "HCBD with this ID already exists"  # Error message
            }
            return render(request, 'AdminaddHCBD.html', dict)

        # Get the foreign key reference for division
        hc_division = Division.objects.get(division_id=hcbdDiv)

        # Create a new Doer entry
        data = HCBD(
            hcbd_id=hcbdID, name=hcbdName, password=hcbdPass, hcbd_email=hcbdEmail, hcbd_division= hc_division)

        # Save the new Doer
        data.save()

        dict = {
            'd_divisionID': d_divisionID,
            'message': "HCBD successfully added"  # Success message
        }
        return render(request, 'AdminaddHCBD.html', dict)

    else:
        dict = {
            'd_divisionID': d_divisionID,
        }
    return render(request, 'AdminaddHCBD.html', dict)

def AdminSearchHCBD(request):
    hcbd_id = request.GET.get('hcbd_id')
    
    # Fetch all hcbd records for the dropdown
    hcbd_list = HCBD.objects.all()

    # Initialize the query to fetch all HCBDs or filter based on the selected HCBD ID
    if hcbd_id:
        hcbds = HCBD.objects.filter(hcbd_id=hcbd_id)
    else:
        hcbds = HCBD.objects.all()  # Show all HCBDs if no filter is applied

    return render(request, 'AdminSearchHCBD.html', {
        'hcbd_list': hcbd_list,  # List for dropdown
        'list_hcbd': hcbds,      # Filtered or all HCBD records to display
    })

def AdminDeleteHCBD(request, hcbd_id):
    hcbd = get_object_or_404(HCBD, hcbd_id=hcbd_id)  
    hcbd.delete()  
    return redirect('AdminSearchHCBD')

def AdminUpdateHCBD(request, hcbd_id):
    hcbd = get_object_or_404(HCBD, hcbd_id=hcbd_id)
    divisions = Division.objects.all()  # Fetch all divisions from the database

    if request.method == 'POST':
        new_division_id = request.POST.get('hcbd_division')
        new_division = Division.objects.get(division_id=new_division_id)  # Fetch the selected division
        hcbd.hcbd_division = new_division
        hcbd.save()

        # Add a success message
        messages.success(request, 'Updated successfully.')

        # Redirect to the same page to prevent form resubmission on refresh
        return redirect('AdminUpdateHCBD', hcbd_id=hcbd_id)

    context = {
        'hcbd_id': hcbd.hcbd_id,
        'name': hcbd.name,
        'password': hcbd.password,
        'hcbd_email': hcbd.hcbd_email,
        'division': hcbd.hcbd_division,
        'divisions': divisions,  # Pass the divisions to the template
    }
    return render(request, 'AdminUpdateHCBD.html', context)

def AdminProfile(request):
    if 'adminID' in request.session:
        adminid = request.session['adminID']
        try:
            admin = Admin.objects.get(admin_id = adminid)
        except Admin.DoesNotExist:
            return redirect('login')
        
        context = {
            'admin' : admin
        }
        return render(request, 'AdminProfile.html', context)
    else:
        return redirect('login')
    

